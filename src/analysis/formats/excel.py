"""Análisis de Excel (XLSX / XLS legado).

Particularidad del catálogo jcyl: las 346 distribuciones declaradas como XLSX
tienen URL con extensión .xls, y el contenido real varía (la mayoría sirve
XLSX real). Por eso el análisis decide el formato por MAGIC BYTES, no por la
extensión:

  - PK\\x03\\x04 (ZIP)      -> XLSX real: openpyxl (vía BytesIO)
  - \\xd0\\xcf\\x11\\xe0 (OLE2) -> Excel 97-2003 legado: xlrd
  - otro                    -> el contenido no es un libro Excel
"""
from __future__ import annotations

import io
from pathlib import Path

from .tabular import _collect_frictionless, _score_from_issues, _normalize, _check_column_quality, _append_quality_issues, _merge_issues, _build_schema_and_sample
from ..checks import missing_dependency_issue
from ..occurrences import new_issue

# Hojas que valida Frictionless. Sigue acotado porque `Resource.validate()`
# reabre y reparsea el libro entero por cada hoja, y el coste es cuadrático en
# libros con decenas de hojas. NO limita las comprobaciones propias, que sí
# recorren el libro completo: por eso se declara en `metrics.sheets_validated`,
# para que la ficha pueda decir sobre qué se ha validado la estructura.
MAX_SHEETS_FRICTIONLESS = 3


def _analyze_with(load_fn, path: Path, ctx: dict) -> dict:
    """Análisis común: hoja(s), validación Frictionless y calidad de columnas.

    `load_fn` recibe la ruta y devuelve un objeto con .sheetnames y .__getitem__.
    """
    issues: list[dict] = []
    issues_sum = 0
    try:
        wb = load_fn()
    except Exception as exc:
        return _normalize(path, ctx, False, 0, "XLSX no válido (no se puede abrir)", {}, [
            {"code": "xlsx-invalido", "label": f"El archivo Excel no se puede abrir: {exc}", "severity": "error", "count": 1},
        ])

    try:
        sheet_names = wb.sheetnames
        sheets = []
        total_rows = 0

        # Validación tabular de las primeras hojas con Frictionless.
        #
        # `sheets_validated` se cuenta, no se estima. Publicaba
        # `min(hojas, MAX_SHEETS_FRICTIONLESS)`, que da por hecho que las tres
        # pasadas salen bien; cuando una hoja revienta las dos veces el `continue`
        # la deja sin validar y la cifra seguía diciendo tres. La métrica está
        # justamente para que la ficha pueda decir sobre qué se ha comprobado.
        sheets_validated = 0
        for name in sheet_names[:MAX_SHEETS_FRICTIONLESS]:
            try:
                from frictionless import Dialect, Resource
                from frictionless.formats.excel.control import ExcelControl

                dialect = Dialect(controls=[ExcelControl(sheet=name)])
                report = Resource(path=str(path), dialect=dialect, format="excel").validate()
            except Exception:
                try:
                    report = Resource(path=str(path), format="excel").validate()
                except Exception:
                    continue
            sheets_validated += 1
            sheet_issues = _collect_frictionless(report)
            issues.extend(sheet_issues)
            issues_sum += sum(i["count"] for i in sheet_issues)

        # Una pasada por hoja: dimensión (max_row/max_column no son fiables en
        # read_only) y análisis propio de calidad sobre TODAS las filas de TODAS
        # las hojas.
        #
        # Antes esto miraba 2.000 filas de las 3 primeras hojas
        # (`MAX_QUALITY_ROWS`, `MAX_SHEETS_VALIDATED`) mientras `total_rows`
        # contaba el libro entero: `metrics.error_cells` y `metrics.total_rows`
        # salían del mismo objeto calculados sobre poblaciones distintas, y un
        # libro de 50.000 filas declaraba las incidencias de las primeras 2.000.
        type_errors, missing_cells = 0, 0
        type_issue = new_issue("error-tipo", "Valores con un tipo distinto al mayoritario de su columna", "error")
        missing_issue = new_issue("celda-faltante", "Celdas vacías en filas con datos", "warning")
        header: list[str] = []
        first_sheet_rows: list[list] = []
        for name in sheet_names:
            ws = wb[name]
            row_count, col_count = 0, 0
            rows: list[list] = []
            for row in ws.iter_rows(values_only=True):
                row_count += 1
                col_count = max(col_count, len(row))
                if any(v is not None and str(v).strip() for v in row):
                    rows.append(list(row))
            sheets.append({"name": name, "rows": row_count, "columns": col_count})
            total_rows += row_count
            if len(rows) > 1:
                sheet_header = [str(h) if h else f"Col {i + 1}" for i, h in enumerate(rows[0])]
                if not header:
                    header = sheet_header
                    first_sheet_rows = [list(r) for r in rows[1:]]
                te, mc, _, _ = _check_column_quality(
                    rows[1:], sheet_header, sheet=name,
                    type_issue=type_issue, missing_issue=missing_issue,
                )
                type_errors += te
                missing_cells += mc
            # La hoja ya está contada: soltar sus filas antes de abrir la
            # siguiente, o un libro grande retiene todas a la vez.
            del rows
        _append_quality_issues(issues, type_issue, missing_issue)
        issues = _merge_issues(issues)

        score, ok = _score_from_issues(issues)
        if total_rows == 0:
            issues.append({"code": "sin-datos", "label": "El libro no contiene filas de datos", "severity": "error", "count": 1})
            score, ok = 0, False

        summary = (
            f"XLSX válido: {len(sheet_names)} hojas, {total_rows:,} filas en total"
            if ok
            else f"XLSX con problemas: {len(sheet_names)} hojas, {total_rows:,} filas, {len(issues)} tipos de incidencia"
        )
        metrics: dict = {
            "sheets": sheets, "sheet_count": len(sheet_names), "total_rows": total_rows,
            "error_cells": issues_sum + type_errors + missing_cells,
            "sheets_validated": sheets_validated,
        }
        if header:
            metrics["header"] = header
        schema: list[dict] = []
        sample_rows_out: list[list] = []
        if header and first_sheet_rows:
            schema, sample_rows_out = _build_schema_and_sample(header, first_sheet_rows)
        return _normalize(path, ctx, ok, score, summary, metrics, issues, schema=schema, sample_rows=sample_rows_out)
    finally:
        try:
            wb.close()
        except Exception:
            pass


def analyze_xlsx(path: Path, ctx: dict) -> dict:
    try:
        import openpyxl
    except ImportError:
        return _normalize(
            path, ctx, False, None,
            "No analizado: falta openpyxl en el entorno de análisis",
            {}, [missing_dependency_issue("openpyxl")],
        )

    raw = path.read_bytes()
    magic = raw[:8]

    # XLSX real (ZIP) aunque la URL diga .xls
    if magic.startswith(b"PK\x03\x04"):
        return _analyze_with(
            lambda: openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True),
            path, ctx,
        )

    # Excel 97-2003 legado (OLE2): se lee con xlrd (viene con frictionless[excel])
    if magic.startswith(b"\xd0\xcf\x11\xe0"):
        return analyze_xlsx_legacy(path, ctx)

    return _normalize(path, ctx, False, 0,
                      "Declarado XLSX pero el contenido no es un libro Excel",
                      {"magic": magic.hex()},
                      [{"code": "formato-no-esperado",
                        "label": "El archivo declarado como XLSX no es un libro Excel (¿HTML de error, PDF, etc.?)",
                        "severity": "error", "count": 1}])


def analyze_xlsx_legacy(path: Path, ctx: dict) -> dict:
    """XLS legado (OLE2) leído con xlrd."""
    # Con el mismo trato que el resto de lectores, y no `import xlrd` a pelo.
    #
    # Sin esto, un entorno sin xlrd lanzaba `ImportError` desde aquí hasta
    # `engine.py`, que lo archiva como `fallo-analizador`: cierto —algo nuestro se
    # rompió— pero inútil para arreglarlo, porque no dice qué falta. Que sea un
    # `dependencia-faltante` explícito es lo que hace que la ficha nombre la
    # librería y que `--check-deps` pueda abortar antes de descargar nada.
    try:
        import xlrd
    except ImportError:
        return _normalize(
            path, ctx, False, None,
            "No analizado: falta xlrd en el entorno de análisis (Excel 97-2003)",
            {}, [missing_dependency_issue("xlrd")],
        )

    try:
        book = xlrd.open_workbook(filename=str(path))
    except Exception as exc:
        return _normalize(path, ctx, False, 0, "XLS legado no válido (no se puede abrir)", {}, [
            {"code": "xls-legado", "label": f"Excel 97-2003 ilegible: {exc}", "severity": "error", "count": 1},
        ])

    sheets = []
    total_rows = 0
    for sh in book.sheets():
        sheets.append({"name": sh.name, "rows": sh.nrows, "columns": sh.ncols})
        total_rows += sh.nrows

    issues: list[dict] = []
    issues_sum = 0

    # Validación estructural con Frictionless (formato xls).
    #
    # UNA sola pasada, y se cuenta como tal. El bucle recorría hasta tres hojas
    # pero validaba `Resource(path, format="xls")` sin decirle cuál —o sea, la
    # primera— y salía con `break` en el primer acierto: siempre una hoja. Aun
    # así `metrics.sheets_validated` publicaba `min(hojas, 3)`, de modo que un
    # libro de diez hojas afirmaba haber validado tres cuando había mirado una.
    # No hay ningún XLS 97-2003 en el catálogo, así que la cifra nunca llegó a
    # publicarse; se corrige porque la métrica existe para que la ficha pueda
    # decir sobre qué se ha comprobado, y para eso tiene que ser verdad.
    sheets_validated = 0
    try:
        from frictionless import Resource

        report = Resource(path=str(path), format="xls").validate()
        sheet_issues = _collect_frictionless(report)
        issues.extend(sheet_issues)
        issues_sum += sum(i["count"] for i in sheet_issues)
        sheets_validated = 1 if book.sheets() else 0
    except Exception:
        pass

    # Comprobaciones propias de calidad de columnas (valores tipados de xlrd)
    type_errors, missing_cells = 0, 0
    type_issue = new_issue("error-tipo", "Valores con un tipo distinto al mayoritario de su columna", "error")
    missing_issue = new_issue("celda-faltante", "Celdas vacías en filas con datos", "warning")
    header: list[str] = []
    first_sheet_rows: list[list] = []
    for sh in book.sheets():
        rows: list[list] = []
        for ri in range(sh.nrows):
            row = [sh.cell_value(ri, ci) for ci in range(sh.ncols)]
            if any(v is not None and str(v).strip() for v in row):
                rows.append(list(row))
        if len(rows) > 1:
            sheet_header = [str(h) if h else f"Col {i + 1}" for i, h in enumerate(rows[0])]
            if not header:
                header = sheet_header
                first_sheet_rows = [list(r) for r in rows[1:]]
            te, mc, _, _ = _check_column_quality(
                rows[1:], sheet_header, sheet=sh.name,
                type_issue=type_issue, missing_issue=missing_issue,
            )
            type_errors += te
            missing_cells += mc
        del rows
    _append_quality_issues(issues, type_issue, missing_issue)
    issues = _merge_issues(issues)

    # El catálogo declara XLSX pero sirve XLS: incoherencia de metadatos
    issues.append({"code": "xls-legado",
                   "label": "Declarado como XLSX en el catálogo, pero el archivo es Excel 97-2003 (.xls)",
                   "severity": "warning", "count": 1})

    score, ok = _score_from_issues(issues)
    if total_rows == 0:
        issues.append({"code": "sin-datos", "label": "El libro no contiene filas de datos", "severity": "error", "count": 1})
        score, ok = 0, False

    summary = (
        f"XLS legado analizado: {len(sheets)} hojas, {total_rows:,} filas"
        if ok
        else f"XLS legado con problemas: {len(sheets)} hojas, {total_rows:,} filas, {len(issues)} tipos de incidencia"
    )
    metrics: dict = {
        "sheets": sheets, "sheet_count": len(sheets), "total_rows": total_rows,
        "error_cells": issues_sum + type_errors + missing_cells,
        "sheets_validated": sheets_validated,
    }
    if header:
        metrics["header"] = header
    schema: list[dict] = []
    sample_rows_out: list[list] = []
    if header and first_sheet_rows:
        schema, sample_rows_out = _build_schema_and_sample(header, first_sheet_rows)
    return _normalize(path, ctx, ok, score, summary, metrics, issues, schema=schema, sample_rows=sample_rows_out)
